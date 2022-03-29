from openpyxl import load_workbook
import pandas as pd
file='C:/Users/BISWAJIT/OneDrive/Desktop/excel/excel_with_blank.xlsx'
wb = load_workbook(filename=file, read_only=False)
for sheet_number in range (0,len(wb.sheetnames)):
    sheet_name_ws=wb.sheetnames[sheet_number]
    df=pd.read_excel(file,sheet_name=str(sheet_name_ws))
    df=pd.DataFrame(df)
    wb.active=sheet_number
    worksheet=wb.active
    column_list=[]
    for cell in worksheet[1]:
        column_list.append(cell.value)
    #print(column_list)
    row_no=1
    row_no_list=[]
    for row in df.itertuples():
        count=1
        row_no=row_no+1
        for column in column_list:
            if str(row[count])=="nan":
                count=count+1
                
                if (count-1)==worksheet.max_column :
                    row_no_list.append(row_no)
    row_no_list.sort(reverse=True)
    #print(row_no_list)
    for col_del in row_no_list:
        worksheet.delete_rows(col_del)
    wb.save(file)



# i=0
# ww=int(worksheet.max_row)
# empty="No"
# while i<ww:
#     if empty=="yes" :
#         pass
#     else:
#         i=i+1
#     row_ind=str(i)
#     cell_indexA="A"+row_ind
#     cell_indexB="B"+row_ind
#     cell_indexC="C"+row_ind
#     cell_indexD="D"+row_ind
#     cell_indexE="E"+row_ind
#     cell_indexF="F"+row_ind
#     cell_indexG="G"+row_ind
#     cell_indexH="H"+row_ind
#     cell_indexI="I"+row_ind
#     cell_indexJ="J"+row_ind
#     cell_indexK="K"+row_ind
                                                         

#     if str(worksheet[cell_indexA].value)=="None":
#         if str(worksheet[cell_indexB].value)=="None":
#              if str(worksheet[cell_indexC].value)=="None":
#                   if str(worksheet[cell_indexD].value)=="None":
#                        if str(worksheet[cell_indexE].value)=="None":
#                             if str(worksheet[cell_indexF].value)=="None":
#                                 if str(worksheet[cell_indexG].value)=="None":
#                                     if str(worksheet[cell_indexH].value)=="None":
#                                         if str(worksheet[cell_indexI].value)=="None":
#                                             if str(worksheet[cell_indexJ].value)=="None":
#                                                 if str(worksheet[cell_indexK].value)=="None":
#                                                     empty="yes"
#                                                     print(ww)
#                                                     worksheet.delete_rows(i)
#                                                     ww=ww-1
#                                                     print(ww)
#                                                     print("deleted row",i)
#                                                     wb.save(file)
                                                    
        
                                              